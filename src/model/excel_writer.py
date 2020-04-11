import openpyxl
from logging import getLogger
logger = getLogger(__name__)


class ExcelWriter:

    def __init__(self,file_name):
        self.file_name = file_name
        self.sheet = None
        self.book = None
        pass

    def __enter__(self):
        self.open()
        return self

    def __exit__(self, type, value, trace):
        self.save()

    def open(self):
        try :
            self.book = openpyxl.load_workbook(self.file_name)
        except:
            self.book = openpyxl.workbook.Workbook()
        try :
            self.sheet = self.book['Sheet']
        except :
            self.book.create_sheet(title="Sheet", index=1)
            self.sheet = self.book['Sheet']

    def writeHeader(self,headers=[]):
        try :
            for r in range(len(headers)):
                self.sheet.cell(row=1, column=(r+1), value=headers[r])
        except :
            pass
        return

    def write(self, rows=[{}]):
        if len(rows) == 0:
            return

        try :
            for r in range(len(rows)):
                r_second = 1
                menu_elements = rows[r].__dict__
                id = menu_elements["id"]
                for k,v in menu_elements.items():
                    self.sheet.cell(row=(id+1), column=(r_second), value=str(v))
                    r_second = r_second + 1
            logger.debug("write %s rows", len(rows))
        except Exception as e:
            logger.error("error has occured during write")

    def save(self):
        self.book.save(self.file_name)

