import openpyxl
import json

book = openpyxl.load_workbook('../Book1.xlsx')
sheets = book['Sheet1']
path_w = '../Book1.json'

for i in range(2, 100):
    id = sheets.cell(row=i, column=1).value
    category = sheets.cell(row=i, column=2).value
    name = sheets.cell(row=i, column=3).value
    price = sheets.cell(row=i, column=4).value
    description = sheets.cell(row=i, column=5).value
    menus = {
        "id": id,
        "category": category, 
        "name": name,
        "price": price,
        "description": description
    }

    with open(path_w, mode='a') as f:
        f.write(json.dumps(menus, sort_keys=False, ensure_ascii=False, indent=4) + ",")
